import numpy as np
import sys
sys.path.append("E:/Project/Python")
from S_matrix.Layer import Layer
from S_matrix.Grating import Sinusoidal,Triangular,Blazed
from S_matrix.Homogeneous_isotropic_matrix import homogeneous_isotropic_matrix
from S_matrix.Layer_mode import layer_mode
from S_matrix.Star import Star
from S_matrix.Build_scatter_side import build_scatter_side
from S_matrix.CalcEffi import calcEffi
from S_matrix.Plot_Effi import Plot_Effi
from openpyxl import load_workbook

def Set_Polarization(thetai,phi,n1,n2,wavelength,pTE,pTM,m,Nx,accuracy,grating,n,Rough=False):
    '''
    thetai:x方向入射角
    phi:y方向入射角
    n1:入射介质折射率
    n2:基底介质折射率
    wavelength:波长
    pTE:TE分量
    pTM:TM分量
    m:截断数
    '''
    Constant={}
    Constant['thetai']=np.radians(thetai)
    Constant['phi']=np.radians(phi)
    Constant['n1']=n1
    Constant['n2']=n2
    Constant['wavelength']=wavelength
    Constant['pTE']=pTE
    Constant['pTM']=pTM
    Constant['k0']=2*np.pi/wavelength
    Constant['n']=n
    n=[0,0,1]
    kinc=Constant['n1']*np.array([np.sin(Constant['thetai'])*np.cos(Constant['phi']),
                                  np.sin(Constant['thetai'])*np.sin(Constant['phi']),np.cos(Constant['thetai'])])
    Constant['kinc']=kinc
    if thetai==0:
        aTE=[0,1,0]
    else:
        aTE=np.cross(n,Constant['kinc'])/np.linalg.norm(np.cross(n,Constant['kinc']))
    aTM=np.cross(Constant['kinc'],aTE)/np.linalg.norm(np.cross(Constant['kinc'],aTE))
    p=np.dot(pTE,aTE)+np.dot(pTM,aTM)
    p=p/np.linalg.norm(p)
    Constant['p']=p
    Constant['n_Tr']=2*m+1#Truncation number
    Constant['period']=grating.T
    Constant['depth_max']=grating.depth
    Constant['mx']=np.arange(-(Constant['n_Tr']//2),Constant['n_Tr']//2+1)
    Constant['my']=np.arange(-(Constant['n_Tr']//2),Constant['n_Tr']//2+1)
    Constant['Nx']=Nx#x方向上的傅里叶快速变换采样数
    Constant['accuracy']=accuracy
    kx=np.diag(kinc[0]-2*np.pi*Constant['mx']/Constant['k0']/Constant['period'])
    kx=kx.astype('complex')
    Constant['kx']=kx
    ky=np.zeros((Constant['n_Tr'],Constant['n_Tr']))
    ky=ky.astype('complex')
    Constant['ky']=ky
    #====================计算反射侧========================
    kzref=np.zeros((Constant['n_Tr'],Constant['n_Tr']),dtype=complex)
    temp=Constant['n1']**2
    for i in range(Constant['n_Tr']):
        kzref[i,i]=np.sqrt(temp-kx[i,i]**2-ky[i,i]**2,dtype=complex)
    Constant['kzref']=kzref
    Ref_mask=np.abs(np.imag(np.diag(Constant['kzref'])))<Constant['accuracy']
    Ref_set=Constant['mx'][Ref_mask]
    Ref_set_ind=np.where(Ref_mask)
    Constant['Ref_set']=Ref_set
    Constant['Ref_set_ind']=Ref_set_ind
    #==================计算透射侧=============================
    kztrn=np.zeros((Constant['n_Tr'],Constant['n_Tr']),dtype=complex)
    temp=Constant['n2']**2
    for i in range(Constant['n_Tr']):
        kztrn[i,i]=np.sqrt(temp-kx[i,i]**2-ky[i,i]**2,dtype=complex)
    Constant['kztrn']=kztrn
    Trn_mask=np.abs(np.imag(np.diag(Constant['kztrn'])))<Constant['accuracy']
    Trn_set=Constant['mx'][Trn_mask]
    Trn_set_ind=np.where(Trn_mask)
    Constant['Trn_set']=Trn_set
    Constant['Trn_set_ind']=Trn_set_ind
    #==========================================================
    LAM,W=homogeneous_isotropic_matrix(1,1,kx,ky)
    temp=int(W.shape[0]/2)
    W0=W[:temp,:temp]
    V0=W[temp:,:temp]
    Constant['W0']=W0
    Constant['V0']=V0
    Constant['W']=W
    if Rough==True:
        h=Roughness(0.32*1e-6,Nx,115)
        Constant['Rough']=h
    return Constant

def Compute(Constant,layers):
    kx=Constant['kx']
    ky=Constant['ky']
    temp=Constant['n_Tr']//2
    Ref_set_ind=Constant['Ref_set_ind']
    Trn_set_ind=Constant['Trn_set_ind']
    S_global=np.block(
        [[np.zeros((4*temp+2,4*temp+2),dtype=complex),np.eye(4*temp+2,dtype=complex)],
         [np.eye(4*temp+2,dtype=complex),np.zeros((4*temp+2,4*temp+2),dtype=complex)]])
    for i in layers[1:-1]:
        S=layer_mode(i,Constant,'FFT')
        S_global=Star(S_global,S)
    Ssub,Wsub,LAMsub=build_scatter_side(Constant['n2']**2,1,np.diag(kx),np.diag(ky),Constant['W'],transmission_side=True)
    Ssub=np.block([[Ssub[0],Ssub[1]],[Ssub[2],Ssub[3]]])
    Sref,Wref,LAMref=build_scatter_side(1,1,np.diag(kx),np.diag(ky),Constant['W'])
    Sref=np.block([[Sref[0],Sref[1]],[Sref[2],Sref[3]]])
    S_global=Star(S_global,Ssub)
    S_global=Star(Sref,S_global)
    R_effi,T_effi=calcEffi(Constant['p'],Constant,S_global)
    Constant['R_effi']=R_effi[Ref_set_ind]
    Constant['T_effi']=T_effi[Trn_set_ind]
    return Constant

def Slice(layers,grating,Constant):
    '''
    n:切片数
    layers:传进仿真层,对中间层进行切片,并返回新的layers函数
    '''
    Constant['n_extra']=0
    n=Constant['n']
    if grating.name!="Rectangular":
        origin_FillFactor=layers[1].fill_factor
        offset=layers[1].offset
        depth=Constant['depth_max']/Constant['n']#切片层的平均厚度
        layer0=layers[0]
        layer_last=layers[-1]
        layer_new=[]
        layer_new.append(layer0)
        if grating.name=="Blazed":
            #=============具体来说,实现效果为将堆叠结构重整为左边对齐的闪耀光栅结构
            for i in range(n):
                fill_factor=(2*i+1)/2/n*origin_FillFactor
                offset=fill_factor/2-origin_FillFactor/2
                layer=Layer(n=Constant['n2'],t=depth,fill_factor=fill_factor,offset=offset)
                layer_new.append(layer)
            layer_new.append(layer_last)
        elif grating.name=="Triangular":
            for i in range(n):
                fill_factor=(2*i+1)/2/n*origin_FillFactor
                offset=-0.5#取0/-0.5都行,即翻转结构
                layer=Layer(n=Constant['n2'],t=depth,fill_factor=fill_factor,offset=offset)
                layer_new.append(layer)
            layer_new.append(layer_last)
        elif grating.name=="Sinusoidal":
            for i in range(n):
                z1=i*depth
                z2=(i+1)*depth
                V=grating.Volume(z1,z2)
                avg_fillfactor=V/depth/grating.T
                layer=Layer(n=Constant['n2'],t=depth,fill_factor=avg_fillfactor,offset=0)
                layer_new.append(layer)
            layer_new.append(layer_last)
    else:
        return layers
    return layer_new

def Roughness(Ra,Nx,seed=None):
    rng=np.random.default_rng(seed)
    h=rng.normal(size=Nx)
    h=h/np.mean(np.abs(h))*Ra
    return h

#============仿真设备层==============================
layers=[
    Layer(n=1,t=1*1e-6),
    Layer(n=1.5,t=2*1e-6,fill_factor=1),
    Layer(n=1.5,t=4*1e-6)
    ]
grating=Sinusoidal(632.8*1e-9*2,1,632.8*1e-9*2)
# grating=Triangular(4*1e-6,30,1)
# grating=Blazed(4*1e-6,30,1,1)
#====================================================
Constant=Set_Polarization(-15,0,1,1.5,632.8*1e-9,1,0,5,2**10,1e-9,grating,n=10,Rough=False)
layers=Slice(layers,grating,Constant)
Constant=Compute(Constant,layers)
Plot_Effi(Constant,[],[])
#========================================================
# file_path='C:/Users/123/Desktop/闪耀光栅30仿真数据.xlsx'
# wb=load_workbook(file_path)
# ws=wb.active
# start_col=2
# start_row=17#行
# counter=0
# save_interval=5
# for n in range(2,101):
#     layers=[
#         Layer(n=1,t=1*1e-6),
#         Layer(n=1.4482+7.5367j,t=2*1e-6,fill_factor=1),
#         Layer(n=1.4482+7.5367j,t=4*1e-6)
#         ]
#     Constant=Set_Polarization(0,0,1,1.5,632.8*1e-9,1,0,5,2**10,1e-9,grating,n)
#     layers=Slice(layers,grating,Constant)
#     Constant=Compute(Constant,layers)
#     R5=Constant['R_effi'][1]
#     ws.cell(row=start_row,column=start_col,value=R5)
#     counter+=1
#     start_row+=1
#     if counter%save_interval==0:
#         wb.save(file_path)
#         print(f"已保存{counter}个数据")
# ws.cell(row=start_row,column=start_col,value="切片数:{},N={}".format(n,101))
# wb.save(file_path)