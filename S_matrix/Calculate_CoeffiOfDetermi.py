import numpy as np
from openpyxl import load_workbook
import re
'''
本脚本用于计算决定系数R2
'''

def Calculate_R2(y_true,y_pred):
    '''
    参数:
    y_ture:真实值数组(观测值)
    y_pred:预测值数组(模型拟合值)
    '''
    #=====转换为numpy数组
    y_true=np.array(y_true)
    y_pred=np.array(y_pred)
    #=====计算观测值的均值
    y_mean=np.mean(y_true)
    #=====计算总平方和
    ss_tot=np.sum((y_true-y_mean)**2)
    #=====计算残差平方和
    ss_res=np.sum((y_true-y_pred)**2)
    #=====计算决定系数
    R2=1-(ss_res/ss_tot)
    return R2

def ReadExcelData(file_path,sheet_name=None,true_col='A',pred_col='B',
                  start_row=2,has_header=True):
    '''
    从Excel文件读取真实值和预测值
    '''
    try:
        wb=load_workbook(filename=file_path,read_only=True)
        if sheet_name is None:
            ws=wb.active
        else:
            ws=wb[sheet_name]
        if has_header:
            start_row=max(start_row,2)
        def col_to_idx(col):
            if isinstance(col,int):
                return col
            elif isinstance(col,str):
                col=col.upper()
                index=0
                for char in col:
                    index=index*26+(ord(char)-ord('A')+1)
                return index
            else:
                raise ValueError(f"无效的列标识:{col}")
        true_col_idx=col_to_idx(true_col)
        pred_col_idx=col_to_idx(pred_col)
        #=====读取数据
        y_true=[]
        y_pred=[]
        row=start_row
        while True:
            true_cell=ws.cell(row=row,column=true_col_idx)
            true_value=true_cell.value

            pred_cell=ws.cell(row=row,column=pred_col_idx)
            pred_value=pred_cell.value
            if true_value is None and pred_value is None:
                break
            if true_value is not None and pred_value is not None:
                try:
                    y_true.append(float(true_value))
                    y_pred.append(float(pred_value))
                except (ValueError,TypeError):
                    print(f"警告:第{row}行数据无法转换为数值,已跳过")
            row+=1
        return np.array(y_true),np.array(y_pred)
    except FileNotFoundError:
        raise FileNotFoundError(f"文件未找到:{file_path}")
    except Exception as e:
        raise Exception(f"读取Excel文件时出错:{str(e)}")

def PercentToFloat(percent_str):
    '''
    使用正则表达将百分数转换为浮点数
    '''
    pass

if __name__=="__main__":
    file_path='C:/Users/123/Desktop/矩形01仿真对比数据.xlsx'
    y_true,y_pred=ReadExcelData(file_path,None,'C','G',2)
    R2=Calculate_R2(y_true,y_pred)
    print(R2)