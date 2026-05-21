import numpy as np
import sys
sys.path.append('E:/Project/python')
from S_matrix.Set_polarization import Set_Polarization
from S_matrix.Layer import Layer
from S_matrix.Compute import Compute
import matplotlib.pyplot as plt
from openpyxl import load_workbook

plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False#解决plt画图中文乱码问题
############################设定仿真设备层#################################
layers=[
    Layer(n=1,t=1*1e-6),
    Layer(n=1.7659,t=2*1e-6,fill_factor=0.5),
    Layer(n=1.7659,t=4*1e-6)
    ]
Constant={}
Constant['n1']=layers[0].n
Constant['e1']=Constant['n1']**2
Constant['n2']=layers[-1].n
Constant['e2']=Constant['n2']**2
###########################设定仿真常数################################
thetai=np.radians(0)#入射角thetai
phi=np.radians(0)#入射角phi
wavelength=632.8*1e-9
pTM=0
pTE=1
Constant=Set_Polarization(thetai,phi,wavelength,pTM,pTE,Constant)
m=30
Constant['n_Tr']=2*m+1
Constant['mx']=np.arange(-(Constant['n_Tr']//2),Constant['n_Tr']//2+1)
Constant['my']=np.arange(-(Constant['n_Tr']//2),Constant['n_Tr']//2+1)
Constant['period']=4*1e-6
Constant['Nx']=2**10
Constant['accuracy']=1e-9
Constant['error']=0.001#相对误差
R_effi=[]
Abs_error=[]
Rela_error=[]
########################数据的输出######################################
# Constant=Compute(Constant,layers,True)
# R_effi=Constant['R_effi']
file_path='C:/Users/123/Desktop/不同材料的0、1级光反射曲线.xlsx'
wb=load_workbook(file_path)
ws=wb.active
start_row=2
start_col=9
for lam in np.arange(300,701,1):
    Constant=Set_Polarization(thetai,phi,lam*1e-9,pTM,pTE,Constant)
    Constant=Compute(Constant,layers,False)
    R_effi=Constant['R_effi']
    temp=np.where(Constant['real_set']==0)[0][0]#0级光在R_effi中的位置
    R1=R_effi[temp+1]
    # R0=R_effi[temp]
    ws.cell(row=start_row,column=start_col,value=R1)
    # ws[f'C{start_row}']=R1
    start_row+=1
wb.save(file_path)