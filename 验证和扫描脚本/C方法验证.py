import matplotlib.pyplot as plt
import numpy as np
import sys
sys.path.append('E:/Project/python')
from C_Method.Grating import Triangular
from C_Method.Compute import Compute
from openpyxl import load_workbook
'''
采用封装式的代码结构
此文件就是main文件
'''

plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False#解决plt画图中文乱码问题

##################设定仿真常数区域##########################################
Constant={}
n1=1
n2=1.02+6.4371j
Constant['n1']=n1
Constant['n2']=n2
Constant['thetai']=np.radians(-10)
Constant['n_Tr']=2*100+1
Constant['wavelength']=632.8*1e-9
Constant['gx']=4*1e-6#结构x方向上的周期
#set Accuracy
Constant['cut']=0#是否对变换后的傅里叶级数进行去除小数处理
Constant['accuracy']=1e-9
Constant['kVecImagMin']=1e-10
R_effi=[]
##########################################################################

##########以下为三角光栅常数设定##############
base_angle=25
grating=Triangular(4*1e-6,base_angle,1)
a,a_diff=grating.profile()
Constant['a']=a#光栅表面轮廓函数
Constant['diff_a']=a_diff#光栅表面轮廓的导数
#############################################

##########任意偏振态,为TE、TM偏振态的组合###############################
alpha=90
alpha=np.radians(alpha)
a=np.cos(alpha)#TM模式的分量
if a<1e-10:
    a=0
b=np.sin(alpha)#TE模式的分量
if b<1e-10:
    b=0
file_path='C:/Users/123/Desktop/44矩阵形式验证.xlsx'
wb=load_workbook(file_path)
ws=wb.active
start_row=2
start_col=6
save_interval=10
counter=0
for lam in range(300,701,1):
    Constant['wavelength']=lam*1e-9
    if a!=0:
        Polarization='TM'
        etaR_TM,etaT_TM=Compute(n1,n2,Polarization,Constant)
    if b!=0:
        Polarization='TE'
        etaR_TE,etaT_TE=Compute(n1,n2,Polarization,Constant)
    if a==0:
        etaR_TM=np.zeros_like(etaR_TE)
    if b==0:
        etaR_TE=np.zeros_like(etaR_TM)
    polar=a**2*etaR_TM+b**2*etaR_TE
    real_Ray1_idx=Constant['real_Ray1_idx']
    temp=np.where(real_Ray1_idx==0)[0][0]#0级光在R_effi中的位置
    try:
        R4=polar[temp-4]
    except IndexError:
        R4=0
    ws.cell(row=start_row,column=start_col,value=R4)
    print("当前波长为{}".format(lam))
    start_row+=1
    counter+=1
    if counter%save_interval==0:
        wb.save(file_path)
        print(f"已保存第{counter}个波长的数据")
wb.save(file_path)
######################################################################
# real_Ray1_idx=Constant['real_Ray1_idx']
# x=np.linspace(min(real_Ray1_idx),max(real_Ray1_idx),max(real_Ray1_idx)-min(real_Ray1_idx)+1,dtype=int)
# plt.plot(x,polar,label='Reflection')
# plt.legend()
# plt.xlabel("Diffraction order")
# plt.ylabel("Diffraction efficiency")
# plt.show()
# print(polar)
# print("sum:"+str(sum(polar)))
print('计算完毕')