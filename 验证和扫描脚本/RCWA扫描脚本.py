import numpy as np
import sys
sys.path.append("E:/Project/Python")
from S_matrix.Layer import Layer
from S_matrix.Grating import Sinusoidal,Triangular,Blazed
from openpyxl import load_workbook
from S_matrix.Set_polarization import Set_Polarization
from S_matrix.Slice import Slice
from S_matrix.Compute import Compute

file_path='C:/Users/123/Desktop/镀膜厚度对反射率的探究.xlsx'
wb=load_workbook(file_path)
ws=wb.active
start_col=2
start_row=2#行
counter=0
save_interval=5
for thick in np.arange(start=1*1e-9,stop=81*1e-9,step=1*1e-9):
    layers=[
        Layer(n=1,t=1*1e-6),
        Layer(n=1.4482+7.5367j,t=2*1e-6,fill_factor=0.9),
        Layer(n=1.4482+7.5367j,t=thick),
        Layer(n=1.457,t=4*1e-6)
        ]
        # grating=Blazed(2*1e-6,angle=angle,fill_factor=fill_factor,n=1)
    grating=Triangular(T=4*1e-6,base_angle=36,fill_factor=0.9)
    Constant=Set_Polarization(thetai=0,phi=0,n1=1,n2=1.457,wavelength=632.8*1e-9,
    pTE=1,pTM=0,m=20,Nx=2**10,accuracy=1e-9,grating=grating,n=50)

    layers=Slice(layers,grating,Constant)
    Constant=Compute(Constant,layers)
    R6=Constant['R_effi'][0]
    ws.cell(row=start_row,column=start_col,value=R6)
    counter+=1
    start_row+=1
    if counter%save_interval==0:
        wb.save(file_path)
        print(f"已保存{counter}个数据")
    # start_col+=1
    # start_row=2
wb.save(file_path)