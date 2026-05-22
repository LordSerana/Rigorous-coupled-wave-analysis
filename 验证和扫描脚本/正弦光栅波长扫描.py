import numpy as np
import sys
sys.path.append("E:/Project/Python")
from S_matrix.Grating import Sinusoidal,Triangular
from C_Method.setConstant import SetConstant
from C_Method.Compute import Compute
import matplotlib.pyplot as plt
from openpyxl import load_workbook

plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False#解决plt画图中文乱码问题

#=======================指定材料等仿真参数=============================
n1=1
n2=1.4482+7.5367j
pol='TE'
n_Tr=2*20+1
lam=632.8*1e-9
thetai=np.radians(0)
cut=0
ImagMin=1e-9
accuracy=1e-10
# grating=Triangular(4*1e-6,36,1)
grating=Sinusoidal(4*1e-6,1,2*1e-6)
Constant=SetConstant(n1,n2,pol,grating.T,n_Tr,lam,thetai)
Constant['depth']=grating.depth
#===================================================================
a=grating.profile()
x=np.linspace(0,Constant['period'],2**10,endpoint=False)
Constant['a']=a(x)#光栅表面轮廓函数
dx=Constant['period']/len(x)
a_diff=np.gradient(Constant['a'],dx)
Constant['a_diff']=a_diff
# plt.plot(x,Constant['a'])
# plt.plot(x,a_diff)
# plt.show()
file_path='C:/Users/123/Desktop/正弦光栅槽深参数扫描.xlsx'
wb=load_workbook(file_path)
ws=wb.active
start_row=2
start_col=2
save_interval=10
counter=0
#=======================================================================
for lam in range(300,700,1):
    Constant=SetConstant(n1,n2,pol,grating.T,n_Tr,lam*1e-9,thetai)
    Constant['depth']=grating.depth
    Constant['a']=a(x)#光栅表面轮廓函数
    Constant['a_diff']=a_diff
    RVec=Compute(Constant)
    R=np.zeros(len(Constant['n1_set']))
    beta1_m=Constant['beta1_m']
    n1_set_ind=Constant['n1_set_ind']
    for i in range(len(Constant['n1_set'])):
        R[i]=np.real(beta1_m[n1_set_ind[i]]/beta1_m[int((Constant['n_Tr']-1)/2)])*(abs(RVec[i])**2)
    T=np.zeros(len(Constant['n1_set']))
    beta2_m=Constant['beta2_m']
    n2_set=Constant['n2_set']
    n2_set_ind=Constant['n2_set_ind']
    if len(n2_set)!=0:
        for i in range(len(n2_set)):
            T[i]=(Constant['eps1']*beta2_m[n2_set_ind[i]]/(Constant['eps2']*beta1_m[(Constant['n_Tr']-1)/2]))*abs(RVec[Constant['n_Tr']+i])**2
    Constant['R_effi']=R
    Constant['T_effi']=T
    temp=np.where(Constant['n1_set']==0)[0][0]
    R0=R[temp]
    ws.cell(row=start_row,column=start_col,value=R0)
    start_row+=1
    counter+=1
    if start_row%save_interval==0:
        wb.save(file_path)
        print(f"已保存{counter}个数据")
wb.save(file_path)